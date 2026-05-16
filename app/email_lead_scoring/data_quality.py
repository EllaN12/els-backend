
"""
Data Quality Report Generator
-------------------------------
Reads a CSV or Excel file and generates a multi-tab Excel data quality report.

Tabs produced:
  1. Summary        — dataset-level overview and quality score
  2. Completeness   — missing value analysis per column
  3. Uniqueness     — duplicate rows and high-cardinality columns
  4. Validity       — data type mismatches, outliers, negative checks
  5. Consistency    — constant columns, mixed types, suspicious patterns
  6. Issues Log     — consolidated list of all flagged issues
"""


import sys
import os
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime


# ── Palette ───────────────────────────────────────────────────────────────────

C_DARK_BLUE  = "1F4E79"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "DEEAF1"
C_WHITE      = "FFFFFF"
C_ALT_ROW    = "F2F7FB"
C_RED        = "C00000"
C_ORANGE     = "ED7D31"
C_GREEN      = "70AD47"
C_YELLOW_BG  = "FFF2CC"
C_RED_BG     = "FCE4D6"
C_GREEN_BG   = "E2EFDA"
C_GREY_BG    = "F2F2F2"

HDR_FONT     = Font(name="Arial", bold=True, color=C_WHITE, size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
TITLE_FONT   = Font(name="Arial", bold=True, size=14, color=C_DARK_BLUE)
SECTION_FONT = Font(name="Arial", bold=True, size=11, color=C_DARK_BLUE)

HDR_FILL     = PatternFill("solid", start_color=C_DARK_BLUE)
SUB_HDR_FILL = PatternFill("solid", start_color=C_MID_BLUE)
ALT_FILL     = PatternFill("solid", start_color=C_ALT_ROW)
WHITE_FILL   = PatternFill("solid", start_color=C_WHITE)
RED_FILL     = PatternFill("solid", start_color=C_RED_BG)
YELLOW_FILL  = PatternFill("solid", start_color=C_YELLOW_BG)
GREEN_FILL   = PatternFill("solid", start_color=C_GREEN_BG)
GREY_FILL    = PatternFill("solid", start_color=C_GREY_BG)

THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_data(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return pd.read_csv(filepath)
    elif ext in (".xlsx", ".xls"):
        return pd.read_excel(filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def style_cell(cell, font=None, fill=None, align=None, border=True, number_format=None):
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = THIN
    if number_format: cell.number_format = number_format


def write_header_row(ws, row_num, headers, col_widths=None):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=c, value=h)
        style_cell(cell, font=HDR_FONT, fill=HDR_FILL,
                   align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[row_num].height = 22
    if col_widths:
        for c, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w


def write_title(ws, row_num, text, ncols):
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num,   end_column=ncols)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 28


def severity_fill(severity):
    return {"High": RED_FILL, "Medium": YELLOW_FILL, "Low": GREEN_FILL}.get(severity, WHITE_FILL)


def severity_font(severity):
    color = {"High": C_RED, "Medium": "BF8F00", "Low": "375623"}.get(severity, "000000")
    return Font(name="Arial", bold=True, size=10, color=color)


# ── Analysis functions ────────────────────────────────────────────────────────

def analyse_completeness(df):
    total = len(df)
    rows = []
    for col in df.columns:
        null_count = df[col].isnull().sum()
        null_pct   = null_count / total * 100 if total > 0 else 0
        if null_pct == 0:
            severity = "Low"
            status   = "✓ Complete"
        elif null_pct <= 5:
            severity = "Low"
            status   = "Minor gaps"
        elif null_pct <= 20:
            severity = "Medium"
            status   = "Moderate gaps"
        else:
            severity = "High"
            status   = "Significant gaps"
        rows.append({
            "Column":        col,
            "Total Rows":    total,
            "Non-Null":      total - null_count,
            "Null Count":    null_count,
            "Null %":        round(null_pct, 2),
            "Completeness %": round(100 - null_pct, 2),
            "Severity":      severity,
            "Status":        status,
        })
    return rows


def analyse_uniqueness(df):
    total = len(df)
    dup_rows = df.duplicated().sum()
    col_rows = []
    for col in df.columns:
        unique   = df[col].nunique()
        dup_vals = total - df[col].nunique(dropna=False)  # rough indicator
        unique_pct = unique / total * 100 if total > 0 else 0
        if unique == 1:
            severity = "High"
            note     = "Constant column — no variance"
        elif unique == total:
            severity = "Low"
            note     = "All values unique (possible ID column)"
        elif unique_pct < 1:
            severity = "Medium"
            note     = "Very low cardinality"
        else:
            severity = "Low"
            note     = ""
        col_rows.append({
            "Column":       col,
            "Unique Values": unique,
            "Unique %":     round(unique_pct, 2),
            "Severity":     severity,
            "Note":         note,
        })
    return dup_rows, col_rows


def analyse_validity(df):
    rows = []
    for col in df.columns:
        series   = df[col].dropna()
        dtype    = str(df[col].dtype)
        issues   = []
        severity = "Low"

        # Numeric outlier check (IQR method)
        if pd.api.types.is_numeric_dtype(df[col]):
            q1, q3 = series.quantile(0.25), series.quantile(0.75)
            iqr    = q3 - q1
            lower, upper = q1 - 3 * iqr, q3 + 3 * iqr
            outliers = ((series < lower) | (series > upper)).sum()
            if outliers > 0:
                issues.append(f"{outliers} outlier(s) (IQR ×3)")
                severity = "Medium"

            # Negative values check
            neg = (series < 0).sum()
            if neg > 0:
                issues.append(f"{neg} negative value(s)")
                severity = max(severity, "Medium", key=["Low","Medium","High"].index)

        # Object column — check for mixed numeric/text
        if dtype == "object":
            numeric_like = pd.to_numeric(series, errors="coerce").notnull().sum()
            if 0 < numeric_like < len(series):
                issues.append(f"Mixed types: {numeric_like} numeric-like in text column")
                severity = "Medium"

            # Check for suspiciously short or empty strings
            empty_str = (series.astype(str).str.strip() == "").sum()
            if empty_str > 0:
                issues.append(f"{empty_str} blank string(s)")
                severity = "Medium"

            # Check for potential date strings stored as text
            date_like = series.astype(str).str.match(
                r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}").sum()
            if date_like > len(series) * 0.5:
                issues.append("Possible date column stored as text")
                severity = max(severity, "Medium", key=["Low","Medium","High"].index)

        rows.append({
            "Column":    col,
            "Data Type": dtype,
            "Issues":    "; ".join(issues) if issues else "None detected",
            "Severity":  severity if issues else "Low",
        })
    return rows


def analyse_consistency(df):
    rows = []
    for col in df.columns:
        series = df[col].dropna()
        issues = []
        severity = "Low"

        # Constant columns
        if df[col].nunique(dropna=False) == 1:
            issues.append("Constant — same value in every row")
            severity = "High"

        # Leading / trailing whitespace in strings
        if str(df[col].dtype) == "object":
            ws_issues = series.astype(str).str.match(r"^\s|\s$").sum()
            if ws_issues > 0:
                issues.append(f"{ws_issues} value(s) with leading/trailing whitespace")
                severity = max(severity, "Medium", key=["Low","Medium","High"].index)

            # Inconsistent casing (mix of upper and lower)
            has_upper = series.astype(str).str.contains(r"[A-Z]").any()
            has_lower = series.astype(str).str.contains(r"[a-z]").any()
            if has_upper and has_lower and series.nunique() > 5:
                issues.append("Mixed casing detected")
                severity = max(severity, "Low", key=["Low","Medium","High"].index)

        rows.append({
            "Column":   col,
            "Issues":   "; ".join(issues) if issues else "None detected",
            "Severity": severity if issues else "Low",
        })
    return rows


def build_issues_log(completeness, validity, consistency):
    log = []
    for r in completeness:
        if r["Null Count"] > 0:
            log.append({
                "Column":   r["Column"],
                "Category": "Completeness",
                "Issue":    f"{r['Null Count']} missing values ({r['Null %']}%)",
                "Severity": r["Severity"],
            })
    for r in validity:
        if r["Issues"] != "None detected":
            log.append({
                "Column":   r["Column"],
                "Category": "Validity",
                "Issue":    r["Issues"],
                "Severity": r["Severity"],
            })
    for r in consistency:
        if r["Issues"] != "None detected":
            log.append({
                "Column":   r["Column"],
                "Category": "Consistency",
                "Issue":    r["Issues"],
                "Severity": r["Severity"],
            })
    # Sort by severity
    order = {"High": 0, "Medium": 1, "Low": 2}
    log.sort(key=lambda x: order.get(x["Severity"], 3))
    return log


def compute_quality_score(df, completeness, dup_rows, issues_log):
    total = len(df)
    # Completeness score
    avg_null_pct = sum(r["Null %"] for r in completeness) / len(completeness) if completeness else 0
    completeness_score = max(0, 100 - avg_null_pct)

    # Uniqueness score (penalise duplicate rows)
    dup_pct = dup_rows / total * 100 if total > 0 else 0
    uniqueness_score = max(0, 100 - dup_pct * 2)

    # Validity / consistency (penalise high/medium issues)
    high   = sum(1 for i in issues_log if i["Severity"] == "High")
    medium = sum(1 for i in issues_log if i["Severity"] == "Medium")
    issue_penalty = min(40, high * 10 + medium * 3)
    validity_score = max(0, 100 - issue_penalty)

    overall = round((completeness_score * 0.4 + uniqueness_score * 0.3 + validity_score * 0.3), 1)
    return {
        "Completeness Score": round(completeness_score, 1),
        "Uniqueness Score":   round(uniqueness_score, 1),
        "Validity Score":     round(validity_score, 1),
        "Overall Score":      overall,
    }


# ── Sheet writers ─────────────────────────────────────────────────────────────

def write_summary_sheet(ws, df, source_file, dup_rows, issues_log, scores):
    total = len(df)
    high   = sum(1 for i in issues_log if i["Severity"] == "High")
    medium = sum(1 for i in issues_log if i["Severity"] == "Medium")
    low    = sum(1 for i in issues_log if i["Severity"] == "Low")
    overall = scores["Overall Score"]

    write_title(ws, 1, "Data Quality Report — Summary", 4)

    ws.cell(row=2, column=1, value=f"Source: {os.path.basename(source_file)}").font = \
        Font(name="Arial", size=10, italic=True, color="595959")
    ws.cell(row=2, column=2, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = \
        Font(name="Arial", size=10, italic=True, color="595959")

    # Dataset stats
    ws.cell(row=4, column=1, value="Dataset Overview").font = SECTION_FONT
    stats = [
        ("Total Rows",         f"{total:,}"),
        ("Total Columns",      len(df.columns)),
        ("Numeric Columns",    sum(pd.api.types.is_numeric_dtype(df[c]) for c in df.columns)),
        ("Text Columns",       sum(str(df[c].dtype) == "object" for c in df.columns)),
        ("DateTime Columns",   sum(pd.api.types.is_datetime64_any_dtype(df[c]) for c in df.columns)),
        ("Duplicate Rows",     f"{dup_rows:,}"),
        ("Total Issues Found", len(issues_log)),
    ]
    for i, (label, value) in enumerate(stats, start=5):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws.cell(row=i, column=2, value=value).font = BODY_FONT
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row=i, column=1).fill = fill
        ws.cell(row=i, column=2).fill = fill
        ws.cell(row=i, column=1).border = THIN
        ws.cell(row=i, column=2).border = THIN

    # Quality scores
    ws.cell(row=13, column=1, value="Quality Scores").font = SECTION_FONT
    score_rows = [
        ("Completeness Score", scores["Completeness Score"]),
        ("Uniqueness Score",   scores["Uniqueness Score"]),
        ("Validity Score",     scores["Validity Score"]),
        ("Overall Score",      scores["Overall Score"]),
    ]
    for i, (label, score) in enumerate(score_rows, start=14):
        ws.cell(row=i, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=i, column=2, value=f"{score}/100")
        if score >= 80:
            cell.font = Font(name="Arial", bold=True, size=10, color="375623")
            cell.fill = GREEN_FILL
        elif score >= 60:
            cell.font = Font(name="Arial", bold=True, size=10, color="BF8F00")
            cell.fill = YELLOW_FILL
        else:
            cell.font = Font(name="Arial", bold=True, size=10, color=C_RED)
            cell.fill = RED_FILL
        ws.cell(row=i, column=1).border = THIN
        cell.border = THIN
        ws.cell(row=i, column=1).fill = ALT_FILL if i % 2 == 0 else WHITE_FILL

    # Issues breakdown
    ws.cell(row=19, column=1, value="Issues by Severity").font = SECTION_FONT
    for i, (label, count, fill) in enumerate([
        ("High",   high,   RED_FILL),
        ("Medium", medium, YELLOW_FILL),
        ("Low",    low,    GREEN_FILL),
    ], start=20):
        ws.cell(row=i, column=1, value=label).font = severity_font(label)
        ws.cell(row=i, column=1).fill = fill
        ws.cell(row=i, column=1).border = THIN
        ws.cell(row=i, column=2, value=count).font = BOLD_FONT
        ws.cell(row=i, column=2).fill = fill
        ws.cell(row=i, column=2).border = THIN

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20


def write_completeness_sheet(ws, rows):
    write_title(ws, 1, "Completeness Analysis", 8)
    headers    = ["Column", "Total Rows", "Non-Null", "Null Count",
                  "Null %", "Completeness %", "Severity", "Status"]
    col_widths = [24, 12, 12, 12, 10, 16, 12, 20]
    write_header_row(ws, 2, headers, col_widths)
    ws.freeze_panes = "A3"

    for i, r in enumerate(rows, start=3):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        values = [r["Column"], r["Total Rows"], r["Non-Null"], r["Null Count"],
                  r["Null %"], r["Completeness %"], r["Severity"], r["Status"]]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font  = BODY_FONT
            cell.fill  = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
        # Colour severity cell
        sev_cell = ws.cell(row=i, column=7)
        sev_cell.fill = severity_fill(r["Severity"])
        sev_cell.font = severity_font(r["Severity"])
        sev_cell.alignment = Alignment(horizontal="center")


def write_uniqueness_sheet(ws, dup_rows, total_rows, col_rows):
    write_title(ws, 1, "Uniqueness Analysis", 5)

    ws.cell(row=2, column=1, value="Duplicate Rows").font = BOLD_FONT
    dup_cell = ws.cell(row=2, column=2,
                       value=f"{dup_rows:,} ({round(dup_rows/total_rows*100,2) if total_rows else 0}%)")
    dup_cell.font = severity_font("High" if dup_rows > 0 else "Low")
    dup_cell.fill = RED_FILL if dup_rows > 0 else GREEN_FILL
    dup_cell.border = THIN
    ws.cell(row=2, column=1).border = THIN

    headers    = ["Column", "Unique Values", "Unique %", "Severity", "Note"]
    col_widths = [24, 14, 12, 12, 40]
    write_header_row(ws, 4, headers, col_widths)
    ws.freeze_panes = "A5"

    for i, r in enumerate(col_rows, start=5):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        values = [r["Column"], r["Unique Values"], r["Unique %"], r["Severity"], r["Note"]]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font  = BODY_FONT
            cell.fill  = fill
            cell.border = THIN
        sev_cell = ws.cell(row=i, column=4)
        sev_cell.fill = severity_fill(r["Severity"])
        sev_cell.font = severity_font(r["Severity"])
        sev_cell.alignment = Alignment(horizontal="center")


def write_validity_sheet(ws, rows):
    write_title(ws, 1, "Validity Analysis", 4)
    headers    = ["Column", "Data Type", "Issues", "Severity"]
    col_widths = [24, 16, 55, 12]
    write_header_row(ws, 2, headers, col_widths)
    ws.freeze_panes = "A3"

    for i, r in enumerate(rows, start=3):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        values = [r["Column"], r["Data Type"], r["Issues"], r["Severity"]]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font  = BODY_FONT
            cell.fill  = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sev_cell = ws.cell(row=i, column=4)
        sev_cell.fill = severity_fill(r["Severity"])
        sev_cell.font = severity_font(r["Severity"])
        sev_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28


def write_consistency_sheet(ws, rows):
    write_title(ws, 1, "Consistency Analysis", 3)
    headers    = ["Column", "Issues", "Severity"]
    col_widths = [24, 60, 12]
    write_header_row(ws, 2, headers, col_widths)
    ws.freeze_panes = "A3"

    for i, r in enumerate(rows, start=3):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        values = [r["Column"], r["Issues"], r["Severity"]]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font  = BODY_FONT
            cell.fill  = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sev_cell = ws.cell(row=i, column=3)
        sev_cell.fill = severity_fill(r["Severity"])
        sev_cell.font = severity_font(r["Severity"])
        sev_cell.alignment = Alignment(horizontal="center")


def write_issues_log_sheet(ws, log):
    write_title(ws, 1, "Issues Log", 4)
    headers    = ["Column", "Category", "Issue", "Severity"]
    col_widths = [24, 16, 60, 12]
    write_header_row(ws, 2, headers, col_widths)
    ws.freeze_panes = "A3"

    if not log:
        ws.cell(row=3, column=1,
                value="✓ No issues detected").font = Font(name="Arial", bold=True,
                                                          size=11, color="375623")
        return

    for i, r in enumerate(log, start=3):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        values = [r["Column"], r["Category"], r["Issue"], r["Severity"]]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font  = BODY_FONT
            cell.fill  = fill
            cell.border = THIN
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sev_cell = ws.cell(row=i, column=4)
        sev_cell.fill = severity_fill(r["Severity"])
        sev_cell.font = severity_font(r["Severity"])
        sev_cell.alignment = Alignment(horizontal="center")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_data_quality_report.py <input_file> [output_file]")
        sys.exit(1)

    input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        base = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{base}_data_quality_report.xlsx"

    print(f"Reading: {input_file}")
    df = load_data(input_file)
    total = len(df)
    print(f"  {total:,} rows × {len(df.columns)} columns")

    print("  Analysing completeness...")
    completeness = analyse_completeness(df)

    print("  Analysing uniqueness...")
    dup_rows, col_uniqueness = analyse_uniqueness(df)

    print("  Analysing validity...")
    validity = analyse_validity(df)

    print("  Analysing consistency...")
    consistency = analyse_consistency(df)

    print("  Building issues log...")
    issues_log = build_issues_log(completeness, validity, consistency)

    scores = compute_quality_score(df, completeness, dup_rows, issues_log)

    print("  Writing Excel report...")
    wb = Workbook()

    ws_summary     = wb.active;           ws_summary.title = "Summary"
    ws_complete    = wb.create_sheet("Completeness")
    ws_unique      = wb.create_sheet("Uniqueness")
    ws_validity    = wb.create_sheet("Validity")
    ws_consistency = wb.create_sheet("Consistency")
    ws_issues      = wb.create_sheet("Issues Log")

    write_summary_sheet(ws_summary,     df, input_file, dup_rows, issues_log, scores)
    write_completeness_sheet(ws_complete, completeness)
    write_uniqueness_sheet(ws_unique,   dup_rows, total, col_uniqueness)
    write_validity_sheet(ws_validity,   validity)
    write_consistency_sheet(ws_consistency, consistency)
    write_issues_log_sheet(ws_issues,   issues_log)

    wb.save(output_file)

    print(f"\nReport saved to: {output_file}")
    print(f"\nQuality Scores:")
    for k, v in scores.items():
        print(f"  {k}: {v}/100")
    print(f"\nTotal issues flagged: {len(issues_log)}")
    high = sum(1 for i in issues_log if i["Severity"] == "High")
    if high:
        print(f"  ⚠  {high} high-severity issue(s) need attention")


if __name__ == "__main__":
    main()

